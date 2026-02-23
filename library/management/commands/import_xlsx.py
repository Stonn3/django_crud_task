from pathlib import Path
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from library.models import Author, Genre, Book

class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument("--file", default="data.xlsx")

    def handle(self, *args, **options):
        p = Path(options["file"]).resolve()
        wb = load_workbook(str(p), data_only=True)

        # authors
        ws = wb["author"]
        rows = list(ws.iter_rows(values_only=True))
        h = rows[0]; idx = {k: h.index(k) for k in h}
        for r in rows[1:]:
            if r[idx["id"]] is None: continue
            Author.objects.update_or_create(
                id=int(r[idx["id"]]),
                defaults={
                    "first_name": r[idx["first_name"]] or "",
                    "last_name": r[idx["last_name"]] or "",
                    "bio": r[idx["bio"]],
                    "birth_date": r[idx["birth_date"]],
                },
            )

        # genres
        ws = wb["genre"]
        rows = list(ws.iter_rows(values_only=True))
        h = rows[0]; idx = {k: h.index(k) for k in h}
        for r in rows[1:]:
            if r[idx["id"]] is None: continue
            Genre.objects.update_or_create(
                id=int(r[idx["id"]]),
                defaults={"name": r[idx["name"]] or "", "description": r[idx["description"]]},
            )

        # books
        ws = wb["book"]
        rows = list(ws.iter_rows(values_only=True))
        h = rows[0]; idx = {k: h.index(k) for k in h}
        for r in rows[1:]:
            if r[idx["id"]] is None: continue
            author = None
            if r[idx["author_id"]] is not None:
                author = Author.objects.filter(id=int(r[idx["author_id"]])).first()

            book, _ = Book.objects.update_or_create(
                id=int(r[idx["id"]]),
                defaults={
                    "title": r[idx["title"]] or "",
                    "author": author,
                    "isbn": str(r[idx["isbn"]]).strip(),
                    "publication_year": int(r[idx["publication_year"]]),
                    "summary": r[idx["summary"]],
                },
            )

            raw = r[idx["genres"]]
            if raw is None:
                ids = []
            elif isinstance(raw, int):
                ids = [raw]
            else:
                ids = [int(x.strip()) for x in str(raw).split(",") if x.strip()]
            book.genres.set(Genre.objects.filter(id__in=ids))

        self.stdout.write("OK")